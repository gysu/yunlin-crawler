import os
import time
import requests
import pandas as pd
from io import StringIO
from dotenv import load_dotenv
import ddddocr
from playwright.sync_api import sync_playwright, Page, expect

# --- 環境設定 ---
load_dotenv()
ACCOUNT = os.getenv('account')
PWD = os.getenv('pwd')
LINE_CHANNEL_ACCESS_TOKEN = os.getenv('LINE_CHANNEL_ACCESS_TOKEN')
LINE_USER_ID = os.getenv('LINE_USER_ID')

# 取得腳本所在的目錄
SCRIPT_DIR = os.path.dirname(os.path.realpath(__file__))
CAPTCHA_IMAGE_PATH = os.path.join(SCRIPT_DIR, "captcha_image.png")

# --- Playwright 瀏覽器操作 ---

def login(page: Page):
    """使用 Playwright 進行登入，包含驗證碼辨識"""
    print("導向登入頁面...")
    # 這裡的 AppIn 是原先登入的入口
    page.goto("https://pwd.yunlin.gov.tw/YLPipe/AppIn")

    print("填寫帳號和密碼...")
    page.locator('input[name="ID"]').fill(ACCOUNT)
    page.locator('input[name="Password"]').fill(PWD)

    print("處理驗證碼...")
    # 找到驗證碼圖片元素並截圖
    captcha_element = page.locator("#CaptchaImg")
    captcha_element.screenshot(path=CAPTCHA_IMAGE_PATH)

    # 使用 ddddocr 進行辨識
    ocr = ddddocr.DdddOcr(beta=True, show_ad=False)
    with open(CAPTCHA_IMAGE_PATH, 'rb') as f:
        image_bytes = f.read()
    captcha_code = ocr.classification(image_bytes)
    print(f"辨識出的驗證碼: {captcha_code}")

    # 填入驗證碼
    page.locator('input[id="Captcha"]').fill(captcha_code)

    print("點擊登入按鈕...")
    page.click('button[onclick="RegCheck();"]')

    # 等待登入成功後的頁面加載
    page.wait_for_url("https://pwd.yunlin.gov.tw/YLPipe/AppIn", timeout=120000)
    print("登入成功！")


# --- 主程式 ---

def main():
    with sync_playwright() as p:
        # 使用 Chrome 瀏覽器，有頭模式方便觀察
        browser = p.chromium.launch(channel='chrome', headless=False, args=["--start-maximized"])
        context = browser.new_context(no_viewport=True)
        page = context.new_page()

        try:
            # 1. 執行登入
            login(page)
            
            # 2. 導向施工打卡頁面
            print("導向施工打卡頁面...")
            page.goto("https://pwd.yunlin.gov.tw/YLPipe/WorkCheckIn")
            
            # 等待網路閒置，確保網頁資料加載完畢
            print("等待資料載入...")
            page.wait_for_load_state('networkidle')
            
            # 取得網頁原始碼
            html_content = page.content()
            
            # 使用 BeautifulSoup 解析 HTML 來精確獲取第二個表格
            from bs4 import BeautifulSoup
            import re
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # 取得 h2 標題作為檔名
            h2_tag = soup.find("h2")
            title_text = h2_tag.text.strip() if h2_tag else "打卡清單"
            # 移除非法字元
            safe_title = re.sub(r'[\\/*?:"<>|]', "", title_text)
            
            # 從標題萃取日期格式 (轉成西元年 YYYYMMDD) 作為下載照片的參數
            date_match = re.search(r'(\d+)年(\d+)月(\d+)日', title_text)
            if date_match:
                roc_year, month, day = date_match.groups()
                gregorian_year = int(roc_year) + 1911
                date_str = f"{gregorian_year}{int(month):02d}{int(day):02d}"
            else:
                from datetime import datetime
                now = datetime.now()
                date_str = f"{now.year}{now.month:02d}{now.day:02d}"
            
            # 根據你提供的 HTML，我們想要的表格在 id="GPC" 的 div 裡面
            gpc_div = soup.find(id="GPC")
            if not gpc_div:
                print("找不到包含資料的目標區塊 (id='GPC')")
                return
                
            table = gpc_div.find("table")
            if not table:
                print("在目標區塊中找不到表格")
                return
                
            # 解析表格資料
            data = []
            rows = table.find("tbody").find_all("tr")
            for row in rows:
                cols = row.find_all("td")
                if len(cols) >= 4:
                    # 判斷進場狀態 (是否有「打卡」按鈕)
                    check_in_btn = cols[0].find("input", {"value": "打卡"})
                    check_in_status = "未打卡" if check_in_btn and "display:none" not in check_in_btn.get("style", "") else "已打卡"
                    
                    # 判斷退場狀態 (是否有「退場打卡」按鈕)
                    check_out_btn = cols[1].find("input", {"value": "退場打卡"})
                    check_out_status = "未退場" if check_out_btn and "display:none" not in check_out_btn.get("style", "") else "已退場"
                    
                    # 如果進場狀態為未打卡，則退場狀態設為空值
                    if check_in_status == "未打卡":
                        check_out_status = ""
                    
                    app_no = cols[2].text.strip()
                    address = cols[3].text.strip()
                    
                    data.append({
                        "案件編號": app_no,
                        "施工地點": address,
                        "進場狀態": check_in_status,
                        "退場狀態": check_out_status,
                        "施工日期": "", # 預設為空，有查詢到才填入
                        "備註": "",
                        "地點": "",
                        "設計圖連結": "",
                        "退場連結": ""
                    })
                    
                    # 如果進場狀態為已打卡
                    if check_in_status == "已打卡":
                        
                        # 1. 抓取設計圖資訊
                        if app_no:
                            design_dir = os.path.join(SCRIPT_DIR, "設計圖資訊")
                            os.makedirs(design_dir, exist_ok=True)
                            
                            design_url = f"https://pwd.yunlin.gov.tw/YLPipe/NewApp/CaseShow?AppNo={app_no}&FolderName=DesignPic"
                            data[-1]["設計圖連結"] = design_url
                            design_pic_paths = [] 
                            try:
                                js_code = """
                                async (url) => {
                                    const response = await fetch(url);
                                    const blob = await response.blob();
                                    return new Promise((resolve, reject) => {
                                        const reader = new FileReader();
                                        reader.onloadend = () => resolve(reader.result);
                                        reader.onerror = reject;
                                        reader.readAsDataURL(blob);
                                    });
                                }
                                """
                                print(f"正在下載設計圖 PDF: {design_url}")
                                b64_data_url = page.evaluate(js_code, design_url)
                                
                                if b64_data_url and b64_data_url.startswith("data:application/pdf;base64,"):
                                    import base64
                                    import fitz  
                                    
                                    b64_data = b64_data_url.split(",")[1]
                                    pdf_bytes = base64.b64decode(b64_data)
                                    
                                    pdf_save_dir = os.path.join(SCRIPT_DIR, "設計圖PDF")
                                    os.makedirs(pdf_save_dir, exist_ok=True)
                                    pdf_file_path = os.path.join(pdf_save_dir, f"{app_no}_design.pdf")
                                    with open(pdf_file_path, "wb") as f:
                                        f.write(pdf_bytes)
                                    print(f"原始設計圖 PDF 已儲存: {pdf_file_path}")
                                    
                                    pdf_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
                                    total_pages = len(pdf_doc)
                                    print(f"偵測到 PDF 共有 {total_pages} 頁")
                                    
                                    for p_num in range(total_pages):
                                        pdf_page = pdf_doc.load_page(p_num)
                                        mat = fitz.Matrix(2.0, 2.0)
                                        pix = pdf_page.get_pixmap(matrix=mat)
                                        
                                        current_path = os.path.join(design_dir, f"{app_no}_design_p{p_num+1}.jpg")
                                        pix.save(current_path)
                                        design_pic_paths.append(current_path)
                                        
                                    pdf_doc.close()
                                    print(f"成功處理設計圖: {app_no}，共轉換 {len(design_pic_paths)} 頁")
                                else:
                                    print(f"下載的不是有效的 PDF 檔案，可能被阻擋: {design_url}")
                                    
                            except Exception as req_e:
                                print(f"處理設計圖時發生錯誤: {design_url}, 錯誤: {req_e}")
                                    
                            data[-1]["設計圖路徑"] = design_pic_paths
                                
                        # 2. 抓取施工進場打卡照片
                        img_dir = os.path.join(SCRIPT_DIR, "施工進場打卡照片")
                        os.makedirs(img_dir, exist_ok=True)
                        
                        img_url = f"https://pwd.yunlin.gov.tw/AppCaseFile/WorkCheckin/{app_no}_{date_str}.jpg"
                        
                        try:
                            img_page = context.new_page()
                            response = img_page.goto(img_url, timeout=15000)
                            
                            if response and response.ok:
                                img_path = os.path.join(img_dir, f"{app_no}_{date_str}.jpg")
                                image_buffer = response.body()
                                with open(img_path, 'wb') as f:
                                    f.write(image_buffer)
                                print(f"成功下載打卡照片: {app_no}_{date_str}.jpg")
                            else:
                                status = response.status if response else "Unknown"
                                print(f"找不到圖片或被阻擋 ({status}): {img_url}")
                        except Exception as req_e:
                            print(f"嘗試載入照片時發生錯誤: {img_url}, 錯誤: {req_e}")
                        finally:
                            if 'img_page' in locals() and not img_page.is_closed():
                                img_page.close()
                                
                        # 3. 查詢並抓取「施工日期」與「備註/地圖」
                        try:
                            # 抓取施工日期
                            query_page = context.new_page()
                            query_page.goto("https://pwd.yunlin.gov.tw/YLPipe/CompleteApp", timeout=15000)
                            query_page.locator('input[name="QueryBase.AppNo"]').fill(app_no)
                            query_page.locator('input[value="查詢"]').click()
                            query_page.wait_for_selector('table[role="grid"]', timeout=15000)
                            
                            query_html = query_page.content()
                            query_soup = BeautifulSoup(query_html, 'html.parser')
                            target_row = query_soup.find("tr", {"data-uid": True})
                            
                            if target_row:
                                cols_query = target_row.find_all("td")
                                if len(cols_query) > 7:
                                    work_date_str = cols_query[7].get_text(separator=" ", strip=True)
                                    print(f"成功查詢到 {app_no} 施工日期: {work_date_str}")
                                    data[-1]["施工日期"] = work_date_str
                            
                            # 抓取備註與地圖資訊
                            query_page.goto("https://pwd.yunlin.gov.tw/YLPub/QueryCases", timeout=15000)
                            query_page.locator('input[name="queryModel.AppNo"]').fill(app_no)
                            query_page.locator('input[value="查詢"]').click()
                            query_page.wait_for_selector('table.table-striped', timeout=15000)
                            
                            pub_html = query_page.content()
                            pub_soup = BeautifulSoup(pub_html, 'html.parser')
                            pub_table = pub_soup.find("table", class_="table-striped")
                            if pub_table:
                                pub_rows = pub_table.find_all("tr")
                                if len(pub_rows) > 1: # 確保有資料列
                                    pub_cols = pub_rows[1].find_all("td")
                                    if len(pub_cols) >= 9:
                                        remark_str = pub_cols[7].get_text(strip=True)
                                        map_link = pub_cols[8].find('a')
                                        map_url = map_link.get('data-src', '') if map_link else ""
                                        
                                        # 從網址擷取經緯度並重新格式化
                                        if map_url:
                                            # 使用正則表達式匹配經緯度數字
                                            coord_match = re.search(r'q=([\d\.]+)\s*,\s*([\d\.]+)', map_url)
                                            if coord_match:
                                                lat, lng = coord_match.groups()
                                                map_url = f"https://www.google.com/maps/search/?api=1&query={lat},{lng}"
                                        
                                        print(f"成功查詢到 {app_no} 備註: {remark_str}")
                                        data[-1]["備註"] = remark_str
                                        data[-1]["地點"] = map_url
                                        
                        except Exception as req_e:
                            print(f"查詢額外資訊時發生錯誤: {app_no}, 錯誤: {req_e}")
                        finally:
                            if 'query_page' in locals() and not query_page.is_closed():
                                query_page.close()

                    # 設定退場照片網址 (無論是否退場都先給予連結)
                    checkout_img_url = f"https://pwd.yunlin.gov.tw/AppCaseFile/WorkCheckout/{app_no}_{date_str}.jpg"
                    data[-1]["退場連結"] = checkout_img_url

                    # 如果退場狀態為已退場，則下載施工退場打卡照片
                    if check_out_status == "已退場":
                        checkout_img_dir = os.path.join(SCRIPT_DIR, "施工退場打卡照片")
                        os.makedirs(checkout_img_dir, exist_ok=True)
                        
                        try:
                            # 建立新分頁
                            checkout_img_page = context.new_page()
                            # 前往圖片網址
                            checkout_response = checkout_img_page.goto(checkout_img_url, timeout=15000)
                            
                            if checkout_response and checkout_response.ok:
                                checkout_img_path = os.path.join(checkout_img_dir, f"{app_no}_{date_str}.jpg")
                                image_buffer = checkout_response.body()
                                with open(checkout_img_path, 'wb') as f:
                                    f.write(image_buffer)
                                print(f"成功下載退場照片: {app_no}_{date_str}.jpg")
                            else:
                                status = checkout_response.status if checkout_response else "Unknown"
                                print(f"找不到退場照片或被阻擋 ({status}): {checkout_img_url}")
                        except Exception as req_e:
                            print(f"嘗試載入退場照片時發生錯誤: {checkout_img_url}, 錯誤: {req_e}")
                        finally:
                            if 'checkout_img_page' in locals() and not checkout_img_page.is_closed():
                                checkout_img_page.close()
            
            # 過濾資料，只保留進場狀態為「已打卡」的項目
            data = [row for row in data if row.get("進場狀態") == "已打卡"]
            
            if data:
                # 確保 DataFrame 欄位順序正確
                df = pd.DataFrame(data)
                cols_order = ["案件編號", "施工地點", "進場狀態", "退場狀態", "施工日期", "備註", "地點", "設計圖連結"]
                # 只保留存在的欄位並照順序排
                df = df[[c for col in cols_order if (c := col) in df.columns]]
                
                # 移除設計圖路徑欄位，避免原始路徑字串出現在 Excel 中
                if "設計圖路徑" in df.columns:
                    df.drop(columns=["設計圖路徑"], inplace=True)
                
                print("--- 成功抓取到打卡表格資料 ---")
                print(df.head())
                
                # 另存為 Excel
                # output_dir = r"E:\gysu_cht\OneDrive - Chunghwa Telecom Co., Ltd\1A_CHT_IMG\crawler_yunlin\0.ar\每日"
                output_dir = os.path.join(SCRIPT_DIR, "output")
                os.makedirs(output_dir, exist_ok=True)
                output_file = os.path.join(output_dir, f"{safe_title}.xlsx")
                
                # 使用 openpyxl 引擎寫入，以支援圖片插入
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='打卡清單')
                    workbook = writer.book
                    worksheet = writer.sheets['打卡清單']
                    
                    from openpyxl.drawing.image import Image
                    from openpyxl.utils import get_column_letter
                    
                    # 重新定義欄位排列:
                    # A: 案件編號, B: 施工地點, C: 進場狀態, D: 退場狀態, E: 施工日期, F: 備註, G: 地點, H: 設計圖連結
                    # I: 進場打卡照片 (第9欄)
                    # J: 退場打卡照片 (第10欄)
                    # K: 設計圖 (第11欄開始)
                    
                    # 設定 I 欄標題 (進場打卡照片)
                    worksheet['I1'] = '進場打卡照片'
                    worksheet.column_dimensions['I'].width = 30
                    
                    # 設定 J 欄標題 (退場打卡照片)
                    worksheet['J1'] = '退場打卡照片'
                    worksheet.column_dimensions['J'].width = 30
                    
                    # 找出所有案件中最多的設計圖頁數
                    max_design_pages = 0
                    for row_data in data:
                        if "設計圖路徑" in row_data:
                            max_design_pages = max(max_design_pages, len(row_data["設計圖路徑"]))
                            
                    # 動態產生設計圖標題 (從 K 欄開始，索引 11)
                    for i in range(max_design_pages):
                        col_letter = get_column_letter(11 + i)
                        worksheet[f'{col_letter}1'] = f'設計圖_第{i+1}頁'
                        worksheet.column_dimensions[col_letter].width = 30
                    
                    for idx, row_data in enumerate(data):
                        excel_row = idx + 2 
                        worksheet.row_dimensions[excel_row].height = 120
                        
                        app_no = row_data["案件編號"]
                        
                        # 設定退場超連結 (D 欄)
                        checkout_url = row_data.get("退場連結", "")
                        if checkout_url:
                            cell = worksheet[f'D{excel_row}']
                            cell.value = row_data["退場狀態"]
                            cell.hyperlink = checkout_url
                            from openpyxl.styles import Font
                            cell.font = Font(color="0000FF", underline="single")

                        # 設定地點超連結 (G 欄)
                        map_url = row_data.get("地點", "")
                        if map_url and map_url.startswith("http"):
                            cell = worksheet[f'G{excel_row}']
                            cell.value = "地點"
                            cell.hyperlink = map_url
                            # 設定藍色字體與底線，讓它看起來像連結
                            from openpyxl.styles import Font
                            cell.font = Font(color="0000FF", underline="single")

                        # 設定設計圖超連結 (H 欄)
                        design_url = row_data.get("設計圖連結", "")
                        if design_url and design_url.startswith("http"):
                            cell = worksheet[f'H{excel_row}']
                            cell.value = "設計圖"
                            cell.hyperlink = design_url
                            from openpyxl.styles import Font
                            cell.font = Font(color="0000FF", underline="single")

                        # 插入進場照片 (I 欄)
                        if row_data["進場狀態"] == "已打卡":
                            img_path = os.path.join(SCRIPT_DIR, "施工進場打卡照片", f"{app_no}_{date_str}.jpg")
                            if os.path.exists(img_path):
                                try:
                                    img = Image(img_path)
                                    img.width = 200
                                    img.height = 150
                                    worksheet.add_image(img, f'I{excel_row}')
                                except Exception as e:
                                    print(f"無法將進場照片加入 Excel: {e}")
                                    
                        # 插入退場照片 (J 欄)
                        if row_data["退場狀態"] == "已退場":
                            checkout_img_path = os.path.join(SCRIPT_DIR, "施工退場打卡照片", f"{app_no}_{date_str}.jpg")
                            if os.path.exists(checkout_img_path):
                                try:
                                    img = Image(checkout_img_path)
                                    img.width = 200
                                    img.height = 150
                                    worksheet.add_image(img, f'J{excel_row}')
                                except Exception as e:
                                    print(f"無法將退場照片加入 Excel: {e}")
                                    
                        # 插入設計圖 (K 欄開始)
                        if "設計圖路徑" in row_data:
                            for i, design_pic_path in enumerate(row_data["設計圖路徑"]):
                                if os.path.exists(design_pic_path):
                                    try:
                                        img = Image(design_pic_path)
                                        img.width = 200
                                        img.height = 150
                                        col_letter = get_column_letter(11 + i)
                                        worksheet.add_image(img, f'{col_letter}{excel_row}')
                                    except Exception as e:
                                        print(f"無法將設計圖加入 Excel: {e}")
                
                print(f"\n已將乾淨的表格資料及照片儲存至: {output_file}")
                
                # --- LINE Messaging API 通知邏輯 ---
                uncheckout_cases = [row["案件編號"] for row in data if row.get("退場狀態") == "未退場"]
                if uncheckout_cases:
                    print("\n檢查到有未退場的案件，準備發送 LINE 通知...")
                    message_text = f"⚠️ 系統提醒 ⚠️\n今日有 {len(uncheckout_cases)} 件案件尚未退場！\n未退場案件編號：\n" + "\n".join(uncheckout_cases)
                    
                    if LINE_CHANNEL_ACCESS_TOKEN and LINE_USER_ID:
                        headers = {
                            "Content-Type": "application/json",
                            "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"
                        }
                        payload = {
                            "to": LINE_USER_ID,
                            "messages": [{"type": "text", "text": message_text}]
                        }
                        try:
                            resp = requests.post("https://api.line.me/v2/bot/message/push", headers=headers, json=payload)
                            if resp.status_code == 200:
                                print("✅ LINE 通知發送成功！")
                            else:
                                print(f"❌ LINE 通知發送失敗: HTTP {resp.status_code}, {resp.text}")
                        except Exception as e:
                            print(f"❌ 發送 LINE 通知時發生錯誤: {e}")
                    else:
                        print("⚠️ 找不到 LINE 的憑證設定，請確認 .env 檔案中是否有 LINE_CHANNEL_ACCESS_TOKEN 和 LINE_USER_ID")
                        
            else:
                print("未能從表格中解析出任何資料。")
            
        except Exception as e:
            print(f"\n腳本執行過程中發生錯誤: {e}")
            page.screenshot(path="error_screenshot.png")
        finally:
            print("關閉瀏覽器...")
            browser.close()

if __name__ == "__main__":
    main()